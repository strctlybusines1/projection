"""
Robust xlsx file reading utilities.

Handles corrupted files, multiple libraries, and edge cases.
"""

import pandas as pd
import warnings
from pathlib import Path
from typing import Optional, List, Dict, Any
import zipfile
import xml.etree.ElementTree as ET


def read_xlsx_robust(filepath: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """
    Robustly read xlsx file trying multiple approaches.

    Attempts:
    1. Standard pandas read_excel with openpyxl
    2. Direct openpyxl with data_only=True
    3. Direct openpyxl with data_only=False (for formula cells)
    4. Manual XML parsing from xlsx zip structure

    Args:
        filepath: Path to xlsx file
        sheet_name: Optional sheet name (uses first/active sheet if None)

    Returns:
        DataFrame with file contents
    """
    filepath = Path(filepath)

    if not filepath.exists():
        raise FileNotFoundError(f"File not found: {filepath}")

    # Suppress warnings during attempts
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")

        # Attempt 1: Standard pandas
        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name or 0, engine='openpyxl')
            if not df.empty:
                return df
        except Exception as e:
            pass

        # Attempt 2: Direct openpyxl with data_only
        try:
            df = _read_with_openpyxl(filepath, sheet_name, data_only=True)
            if df is not None and not df.empty:
                return df
        except Exception as e:
            pass

        # Attempt 3: Direct openpyxl without data_only
        try:
            df = _read_with_openpyxl(filepath, sheet_name, data_only=False)
            if df is not None and not df.empty:
                return df
        except Exception as e:
            pass

        # Attempt 4: Manual XML parsing
        try:
            df = _read_xlsx_as_xml(filepath, sheet_name)
            if df is not None and not df.empty:
                return df
        except Exception as e:
            pass

    raise ValueError(f"Could not read xlsx file with any method: {filepath}")


def _read_with_openpyxl(filepath: Path, sheet_name: Optional[str],
                         data_only: bool = True) -> Optional[pd.DataFrame]:
    """Read xlsx using openpyxl directly."""
    from openpyxl import load_workbook

    wb = load_workbook(filepath, data_only=data_only, read_only=True)

    if not wb.sheetnames:
        # Try without read_only mode
        wb = load_workbook(filepath, data_only=data_only, read_only=False)

    if not wb.sheetnames:
        return None

    # Get target sheet
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active or wb[wb.sheetnames[0]]

    # Extract data
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)

    if not data:
        return None

    # First row as header
    df = pd.DataFrame(data[1:], columns=data[0])
    return df


def _read_xlsx_as_xml(filepath: Path, sheet_name: Optional[str] = None) -> Optional[pd.DataFrame]:
    """
    Read xlsx by parsing XML directly from zip structure.

    xlsx files are zip archives containing XML files.
    Handles both standard and strict OOXML namespaces.
    """
    # Both possible namespaces (standard and strict OOXML)
    NAMESPACES = [
        'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
        'http://purl.oclc.org/ooxml/spreadsheetml/main'
    ]

    with zipfile.ZipFile(filepath, 'r') as zf:
        # List contents
        names = zf.namelist()

        # Find sheet files
        sheet_files = sorted([n for n in names if n.startswith('xl/worksheets/sheet')])

        if not sheet_files:
            return None

        # Get shared strings (for text values) - try both namespaces
        shared_strings = []
        if 'xl/sharedStrings.xml' in names:
            with zf.open('xl/sharedStrings.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()

                for namespace in NAMESPACES:
                    ns = {'ns': namespace}
                    items = root.findall('.//ns:si', ns)
                    if items:
                        for si in items:
                            t = si.find('.//ns:t', ns)
                            shared_strings.append(t.text if t is not None else '')
                        break

        # Read target sheet (default to first, or second if sheet_name suggests 'Projected')
        sheet_idx = 0
        if sheet_name:
            sheet_name_lower = sheet_name.lower()
            if 'proj' in sheet_name_lower and len(sheet_files) > 1:
                sheet_idx = 1
            elif 'actual' in sheet_name_lower:
                sheet_idx = 0

        sheet_file = sheet_files[min(sheet_idx, len(sheet_files) - 1)]

        with zf.open(sheet_file) as f:
            tree = ET.parse(f)
            root = tree.getroot()

            # Try both namespaces
            rows_data = []
            for namespace in NAMESPACES:
                ns = {'ns': namespace}
                rows = root.findall('.//ns:row', ns)
                if rows:
                    for row in rows:
                        row_data = {}
                        for cell in row.findall('ns:c', ns):
                            cell_ref = cell.get('r', '')
                            # Extract column letters (handle multi-letter columns like AA, AB)
                            col = ''.join(c for c in cell_ref if c.isalpha())
                            cell_type = cell.get('t', '')
                            value_elem = cell.find('ns:v', ns)

                            if value_elem is not None:
                                value = value_elem.text
                                # Handle shared strings
                                if cell_type == 's' and value and shared_strings:
                                    try:
                                        idx = int(value)
                                        if idx < len(shared_strings):
                                            value = shared_strings[idx]
                                    except ValueError:
                                        pass
                                row_data[col] = value
                            else:
                                row_data[col] = None

                        if row_data:
                            rows_data.append(row_data)
                    break  # Found data with this namespace

            if not rows_data:
                return None

            # Convert to DataFrame
            df = pd.DataFrame(rows_data)

            # Use first row as header if it looks like headers
            if len(df) > 1:
                df.columns = df.iloc[0]
                df = df.iloc[1:].reset_index(drop=True)

            return df

    return None


def get_xlsx_info(filepath: str) -> Dict[str, Any]:
    """
    Get information about an xlsx file without fully reading it.

    Returns:
        Dict with sheet names, row counts, column info
    """
    filepath = Path(filepath)
    info = {
        'filepath': str(filepath),
        'exists': filepath.exists(),
        'size_bytes': filepath.stat().st_size if filepath.exists() else 0,
        'sheets': [],
        'error': None
    }

    if not filepath.exists():
        info['error'] = 'File not found'
        return info

    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True, data_only=True)
            info['sheets'] = wb.sheetnames

            if wb.sheetnames:
                ws = wb.active or wb[wb.sheetnames[0]]
                info['dimensions'] = ws.dimensions
    except Exception as e:
        info['error'] = str(e)

        # Try zip inspection
        try:
            with zipfile.ZipFile(filepath, 'r') as zf:
                info['is_valid_zip'] = True
                info['zip_contents'] = zf.namelist()
        except:
            info['is_valid_zip'] = False

    return info


def list_sheets(filepath: str) -> List[str]:
    """Get list of sheet names in xlsx file."""
    info = get_xlsx_info(filepath)
    return info.get('sheets', [])


def read_all_sheets_xml(filepath: str) -> Dict[str, pd.DataFrame]:
    """
    Read all sheets from xlsx using XML parsing.

    Handles both standard and strict OOXML namespaces.

    Returns:
        Dict mapping sheet index/name to DataFrame
    """
    filepath = Path(filepath)
    NAMESPACES = [
        'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
        'http://purl.oclc.org/ooxml/spreadsheetml/main'
    ]

    results = {}

    with zipfile.ZipFile(filepath, 'r') as zf:
        names = zf.namelist()
        sheet_files = sorted([n for n in names if n.startswith('xl/worksheets/sheet')])

        # Get shared strings
        shared_strings = []
        if 'xl/sharedStrings.xml' in names:
            with zf.open('xl/sharedStrings.xml') as f:
                tree = ET.parse(f)
                root = tree.getroot()
                for namespace in NAMESPACES:
                    ns = {'ns': namespace}
                    items = root.findall('.//ns:si', ns)
                    if items:
                        for si in items:
                            t = si.find('.//ns:t', ns)
                            shared_strings.append(t.text if t is not None else '')
                        break

        # Read each sheet
        for i, sheet_file in enumerate(sheet_files):
            with zf.open(sheet_file) as f:
                tree = ET.parse(f)
                root = tree.getroot()

                for namespace in NAMESPACES:
                    ns = {'ns': namespace}
                    rows = root.findall('.//ns:row', ns)
                    if rows:
                        rows_data = []
                        for row in rows:
                            row_data = {}
                            for cell in row.findall('ns:c', ns):
                                cell_ref = cell.get('r', '')
                                col = ''.join(c for c in cell_ref if c.isalpha())
                                cell_type = cell.get('t', '')
                                value_elem = cell.find('ns:v', ns)

                                if value_elem is not None:
                                    value = value_elem.text
                                    if cell_type == 's' and value and shared_strings:
                                        try:
                                            idx = int(value)
                                            if idx < len(shared_strings):
                                                value = shared_strings[idx]
                                        except ValueError:
                                            pass
                                    row_data[col] = value

                            if row_data:
                                rows_data.append(row_data)

                        if rows_data:
                            df = pd.DataFrame(rows_data)
                            if len(df) > 1:
                                df.columns = df.iloc[0]
                                df = df.iloc[1:].reset_index(drop=True)
                            results[f'Sheet{i+1}'] = df
                        break

    return results
